#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel错误标注生成器
将验证错误直接标注到原始用户上传的Excel文件中
"""

import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Dict, List
from models.bom_models import ValidationResult, ValidationError
from utils.logger import get_default_logger

logger = get_default_logger()


class ExcelMarkupGenerator:
    """Excel错误标注生成器 - 直接在原文件上标注错误"""
    
    # 颜色定义
    ERROR_FILL = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")  # 浅红色
    WARNING_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # 浅黄色
    ERROR_FONT = Font(color="FF0000", bold=True)  # 红色粗体
    WARNING_FONT = Font(color="FF8800", bold=True)  # 橙色粗体
    
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self):
        self.error_column_offset = 2  # 错误信息列距离最后一列的距离
    
    def generate(self, result: ValidationResult, original_file_path: str, output_path: str):
        """
        在原始Excel文件上标注错误
        
        参数:
        - result: ValidationResult对象，包含验证结果
        - original_file_path: 原始上传的Excel文件路径
        - output_path: 输出文件路径
        """
        logger.info(f"开始生成标注后的Excel文件: {output_path}")
        
        try:
            # 复制原始文件到输出路径
            shutil.copy(original_file_path, output_path)
            logger.info(f"已复制原文件到: {output_path}")
            
            # 加载工作簿
            wb = load_workbook(output_path)
            ws = wb.active
            
            logger.info(f"开始标注错误...")
            
            # 创建行号到错误库的映射
            errors_by_row = self._group_errors_by_row(result)
            
            # 确定表头行数和数据开始行
            header_row = 1
            data_start_row = 2
            
            # 添加错误列标题
            max_col = ws.max_column
            error_column = max_col + 1
            warning_column = max_col + 2
            
            # 添加列标题
            ws.cell(row=header_row, column=error_column).value = "❌错误"
            ws.cell(row=header_row, column=warning_column).value = "⚠️警告"
            
            # 格式化标题
            for col in [error_column, warning_column]:
                cell = ws.cell(row=header_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # 读取表头列名到位置映射
            header_cells = {ws.cell(row=header_row, column=col).value: col for col in range(1, ws.max_column + 1)}

            # 遍历数据行，添加错误标注
            for row_idx, (row_num, errors) in enumerate(errors_by_row.items()):
                # Excel行号 = 数据行号 + 1（表头）
                excel_row = row_num
                
                # 分离错误和警告
                err_list = [e for e in errors if e.severity == 'error']
                warn_list = [e for e in errors if e.severity == 'warning']
                
                # 标注错误
                if err_list:
                    error_text = self._format_errors(err_list)
                    error_cell = ws.cell(row=excel_row, column=error_column)
                    error_cell.value = error_text
                    error_cell.fill = self.ERROR_FILL
                    error_cell.font = self.ERROR_FONT
                    error_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    error_cell.border = self.THIN_BORDER
                
                # 标注警告
                if warn_list:
                    warning_text = self._format_errors(warn_list)
                    warning_cell = ws.cell(row=excel_row, column=warning_column)
                    warning_cell.value = warning_text
                    warning_cell.fill = self.WARNING_FILL
                    warning_cell.font = self.WARNING_FONT
                    warning_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    warning_cell.border = self.THIN_BORDER
                
                # 标记不符合需求的关键字段单元格
                self._highlight_issue_cells(ws, excel_row, errors, header_cells)

                # 标记原数据行（如果有错误则整行浅红）
                self._highlight_row(ws, excel_row, len(err_list) > 0)
            
            # 调整列宽
            ws.column_dimensions['A'].width = 15
            for col in range(max_col + 1, max_col + 3):
                ws.column_dimensions[chr(64 + col) if col <= 26 else chr(64 + (col-26)) + chr(65 + ((col-27) % 26))].width = 40
            
            # 保存工作簿
            wb.save(output_path)
            logger.info(f"标注完成，文件已保存: {output_path}")
            
            # 打印统计信息
            logger.info(f"总计标注了 {len(errors_by_row)} 行数据")
            logger.info(f"  错误行数: {len([e for errors in errors_by_row.values() for e in errors if e.severity == 'error'])//max(1, len(errors_by_row))}")
            logger.info(f"  警告行数: {len([e for errors in errors_by_row.values() for e in errors if e.severity == 'warning'])//max(1, len(errors_by_row))}")
            
        except Exception as e:
            logger.error(f"生成标注Excel文件失败: {str(e)}")
            raise
    
    def _group_errors_by_row(self, result: ValidationResult) -> Dict[int, List[ValidationError]]:
        """按行号分组错误和警告"""
        errors_by_row = {}
        
        # 合并错误和警告
        all_issues = result.errors + result.warnings
        
        for error in all_issues:
            row_num = error.row_number
            if row_num not in errors_by_row:
                errors_by_row[row_num] = []
            errors_by_row[row_num].append(error)
        
        return errors_by_row
    
    def _format_errors(self, errors: List[ValidationError]) -> str:
        """格式化错误信息为文本"""
        formatted = []
        for i, error in enumerate(errors, 1):
            # 格式: [规则ID] 错误信息
            msg = f"[{error.rule_id}] {error.message}"
            
            if error.expected_value and error.actual_value:
                msg += f"\n  期望: {error.expected_value}"
                msg += f"\n  实际: {error.actual_value}"
            
            formatted.append(msg)
        
        return "\n\n".join(formatted)
    
    def _highlight_issue_cells(self, ws, row_num: int, issues: List[ValidationError], header_cells: Dict[str, int]):
        """根据字段在行内标记具体不符合项"""
        field_mapping = {
            'parent_code': '父编码',
            'child_code': '子编码',
            'quantity': '用量',
            'unit': '单位',
            'position_number': '位置号',
            'bomviewaltsuid': 'bomviewaltsuid',
            'order_type': '单别',
            'work_order_category': '工单类别'
        }

        for issue in issues:
            target_fields = list(issue.highlight_fields) if issue.highlight_fields else []
            if issue.field and issue.field not in target_fields:
                target_fields.insert(0, issue.field)

            for field_name in target_fields:
                col = header_cells.get(str(field_name))
                if col is None:
                    mapped = field_mapping.get(str(field_name))
                    if mapped:
                        col = header_cells.get(mapped)

                if col:
                    cell = ws.cell(row=row_num, column=col)
                    if issue.severity == 'error':
                        cell.fill = self.ERROR_FILL
                        cell.font = self.ERROR_FONT
                    else:
                        cell.fill = self.WARNING_FILL
                        cell.font = self.WARNING_FONT
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def _highlight_row(self, ws, row_num: int, has_error: bool):
        """为行添加背景高亮（轻微）"""
        color = "FFEEEE" if has_error else "FFCCFF"  # 错误行浅红，警告行浅紫
        
        for cell in ws[row_num]:
            if cell.value is not None and (not cell.fill or cell.fill.start_color.index == "00000000"):
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


class ExcelErrorReportGenerator:
    """Excel错误详情报告生成器 - 生成详细的错误分析表"""
    
    def __init__(self):
        pass
    
    def generate(self, result: ValidationResult, output_path: str):
        """生成详细的错误分析Excel报告"""
        logger.info(f"开始生成详细错误报告: {output_path}")
        
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "错误详情"
            
            # 写入表头
            headers = ["行号", "字段", "规则ID", "规则名称", "严重程度", "错误信息", "期望值", "实际值"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 写入错误信息
            row_num = 2
            for error in result.errors + result.warnings:
                ws.cell(row=row_num, column=1).value = error.row_number
                ws.cell(row=row_num, column=2).value = error.field
                ws.cell(row=row_num, column=3).value = error.rule_id
                ws.cell(row=row_num, column=4).value = error.rule_name
                ws.cell(row=row_num, column=5).value = error.severity
                ws.cell(row=row_num, column=6).value = error.message
                ws.cell(row=row_num, column=7).value = str(error.expected_value) if error.expected_value else ""
                ws.cell(row=row_num, column=8).value = str(error.actual_value) if error.actual_value else ""
                
                # 按严重程度着色
                fill_color = "FFD9D9" if error.severity == 'error' else "FFFFCC"
                for col in range(1, 9):
                    ws.cell(row=row_num, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                row_num += 1
            
            # 调整列宽
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 30
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 20
            
            # 添加统计sheet
            summary_ws = wb.create_sheet("统计")
            
            summary_ws['A1'] = "校验统计"
            summary_ws['A1'].font = Font(bold=True, size=14)
            
            summary_ws['A3'] = "总行数:"
            summary_ws['B3'] = result.total_rows
            summary_ws['A4'] = "有效行数:"
            summary_ws['B4'] = result.valid_rows
            summary_ws['A5'] = "错误数量:"
            summary_ws['B5'].value = result.error_count
            summary_ws['B5'].fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
            summary_ws['A6'] = "警告数量:"
            summary_ws['B6'].value = result.warning_count
            summary_ws['B6'].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            summary_ws['A7'] = "通过率:"
            summary_ws['B7'] = f"{result.pass_rate:.2f}%"
            
            summary_ws.column_dimensions['A'].width = 15
            summary_ws.column_dimensions['B'].width = 20
            
            wb.save(output_path)
            logger.info(f"详细错误报告生成成功: {output_path}")
            
        except Exception as e:
            logger.error(f"生成错误详情报告失败: {str(e)}")
            raise
