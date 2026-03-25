import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from models.bom_models import ValidationResult
from utils.logger import get_default_logger

logger = get_default_logger()


class ExcelReportGenerator:
    """Excel报告生成器"""
    
    def generate(self, result: ValidationResult, output_path: str):
        """生成Excel报告"""
        logger.info(f"开始生成Excel报告: {output_path}")
        
        try:
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                self._write_summary_sheet(writer, result)
                
                if result.error_count > 0:
                    self._write_errors_sheet(writer, result)
                
                if result.warning_count > 0:
                    self._write_warnings_sheet(writer, result)
                
                self._write_details_sheet(writer, result)
            
            self._apply_formatting(output_path)
            
            logger.info(f"Excel报告生成成功: {output_path}")
            
        except Exception as e:
            logger.error(f"生成Excel报告失败: {str(e)}")
            raise
    
    def _write_summary_sheet(self, writer: pd.ExcelWriter, result: ValidationResult):
        """写入摘要sheet"""
        summary_data = {
            '项目': [
                '文件名称',
                '校验时间',
                '总行数',
                '有效行数',
                '错误数量',
                '警告数量',
                '通过率',
                '校验结果'
            ],
            '值': [
                result.file_name,
                result.validation_time.strftime('%Y-%m-%d %H:%M:%S'),
                result.total_rows,
                result.valid_rows,
                result.error_count,
                result.warning_count,
                f"{result.pass_rate:.2f}%",
                '通过' if result.is_valid else '失败'
            ]
        }
        
        df = pd.DataFrame(summary_data)
        df.to_excel(writer, sheet_name='摘要', index=False)
    
    def _write_errors_sheet(self, writer: pd.ExcelWriter, result: ValidationResult):
        """写入错误sheet"""
        error_data = []
        for error in result.errors:
            error_data.append({
                '行号': error.row_number,
                '规则ID': error.rule_id,
                '规则名称': error.rule_name,
                '严重程度': error.severity,
                '字段': error.field,
                '错误描述': error.message,
                '期望值': error.expected_value if error.expected_value else '',
                '实际值': error.actual_value if error.actual_value else ''
            })
        
        df = pd.DataFrame(error_data)
        df.to_excel(writer, sheet_name='错误列表', index=False)
    
    def _write_warnings_sheet(self, writer: pd.ExcelWriter, result: ValidationResult):
        """写入警告sheet"""
        warning_data = []
        for warning in result.warnings:
            warning_data.append({
                '行号': warning.row_number,
                '规则ID': warning.rule_id,
                '规则名称': warning.rule_name,
                '严重程度': warning.severity,
                '字段': warning.field,
                '警告描述': warning.message,
                '期望值': warning.expected_value if warning.expected_value else '',
                '实际值': warning.actual_value if warning.actual_value else ''
            })
        
        df = pd.DataFrame(warning_data)
        df.to_excel(writer, sheet_name='警告列表', index=False)
    
    def _write_details_sheet(self, writer: pd.ExcelWriter, result: ValidationResult):
        """写入详情sheet"""
        detail_data = []
        for item in result.items:
            errors_for_row = [e for e in result.errors if e.row_number == item.row_number]
            warnings_for_row = [w for w in result.warnings if w.row_number == item.row_number]
            
            status = '通过'
            if errors_for_row:
                status = '错误'
            elif warnings_for_row:
                status = '警告'
            
            detail_data.append({
                '行号': item.row_number,
                '状态': status,
                '物料编码': item.material_code,
                '物料名称': item.material_name,
                '规格型号': item.specification if item.specification else '',
                '数量': item.quantity if item.quantity else '',
                '单位': item.unit if item.unit else '',
                '供应商': item.supplier if item.supplier else '',
                '问题描述': '; '.join([e.message for e in errors_for_row + warnings_for_row])
            })
        
        df = pd.DataFrame(detail_data)
        df.to_excel(writer, sheet_name='详细信息', index=False)
    
    def _apply_formatting(self, file_path: str):
        """应用Excel格式"""
        try:
            wb = load_workbook(file_path)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            
        except Exception as e:
            logger.warning(f"应用Excel格式失败: {str(e)}")
