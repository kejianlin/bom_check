"""
BOM模板生成器
根据配置自动生成标准的BOM Excel模板
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from pathlib import Path
import sys

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))


def create_bom_template(output_path: str = None, num_rows: int = 20):
    """
    创建BOM申请模板
    
    Args:
        output_path: 输出文件路径，如果为None则使用默认路径
        num_rows: 模板中的数据行数（默认20行）
    """
    
    if output_path is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = project_root / 'templates' / 'generated' / f'BOM申请模板_{timestamp}.xlsx'
    
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    print("=" * 80)
    print("BOM模板生成器")
    print("=" * 80)
    print()
    
    # 定义列名（标准化的列名）
    columns = {
        'parent_code': '父编码',
        'bomviewaltsuid': 'bomviewaltsuid',
        'child_code': '子编码',
        'quantity': '用量',
        'position_number': '位置号',
        'remark': '备注',
        'order_type': '单别',
        'work_order_category': '工单类别',
        'unit': '单位'
    }
    
    # 创建空数据框
    df = pd.DataFrame(columns=list(columns.values()))
    
    # 添加示例数据行
    sample_data = {
        '父编码': 'MSP-000163-00',
        'bomviewaltsuid': '0',
        '子编码': 'WIR-001952-00',
        '用量': 1,
        '位置号': '',
        '备注': '',
        '单别': 'BOM清单|BM11',
        '工单类别': '5101',
        '单位': '个'
    }
    
    # 添加一行示例数据
    df = pd.concat([df, pd.DataFrame([sample_data])], ignore_index=True)
    
    # 添加空行
    for _ in range(num_rows - 1):
        empty_row = {col: '' for col in df.columns}
        df = pd.concat([df, pd.DataFrame([empty_row])], ignore_index=True)
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM数据"
    
    # 定义样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='微软雅黑', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center')
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # 简洁的列名（第一行）
    headers = list(columns.values())
    
    # 详细说明（第二行）
    header_notes = [
        '格式: XXX-000000-00',
        '必须填0',
        '格式: XXX-000000-00',
        '数字，最多4位小数',
        '可选',
        '可选',
        'BOM清单|BM11、配电BOM|PBOM、新能源BOM|XBOM、易立高BOM|YBOM',
        '5101-厂内、5102-试产、5103-委外',
        '个、套、米等'
    ]
    
    # 写入第一行表头（列名）
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 写入第二行（说明）
    note_font = Font(name='微软雅黑', size=9, italic=True)
    note_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    note_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_idx, note in enumerate(header_notes, 1):
        cell = ws.cell(row=2, column=col_idx, value=note)
        cell.font = note_font
        cell.fill = note_fill
        cell.alignment = note_alignment
        cell.border = border
    
    # 设置行高
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 40
    
    # 写入数据（从第3行开始）
    for row_idx, row_data in enumerate(df.values, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
            
            # 第一行数据（示例行）使用浅蓝色背景
            if row_idx == 3:
                cell.fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
    
    # 设置列宽
    column_widths = {
        1: 18,  # 父编码
        2: 15,  # bomviewaltsuid
        3: 18,  # 子编码
        4: 10,  # 用量
        5: 12,  # 位置号
        6: 15,  # 备注
        7: 25,  # 单别
        8: 25,  # 工单类别
        9: 10   # 单位
    }
    
    for col_idx, width in column_widths.items():
        ws.column_dimensions[chr(64 + col_idx)].width = width
    
    # 添加说明工作表
    ws_info = wb.create_sheet("填写说明")
    
    info_data = [
        ["BOM申请模板填写说明", ""],
        ["", ""],
        ["字段名称", "填写要求"],
        ["父编码", "格式：3位大写字母-6位数字-2位数字（例如：MSP-000163-00）"],
        ["bomviewaltsuid", "必须填写：0"],
        ["子编码", "格式：3位大写字母-6位数字-2位数字（例如：WIR-001952-00）"],
        ["用量", "数字，支持小数，最多4位小数（例如：1、2.5、0.0001）"],
        ["位置号", "可选，1-10位字母或数字"],
        ["备注", "可选，任意文本"],
        ["单别", "必须是以下之一：\nBOM清单|BM11\n配电BOM|PBOM\n新能源BOM|XBOM\n易立高BOM|YBOM"],
        ["工单类别", "必须是以下之一：\n5101（厂内工单）\n5102（试产工单）\n5103（委外工单）"],
        ["单位", "必须是系统允许的单位（例如：个、套、米、千克等）"],
        ["", ""],
        ["注意事项", ""],
        ["1", "所有必填字段不能为空"],
        ["2", "编码格式必须严格按照要求填写"],
        ["3", "单别和工单类别必须从允许的值中选择"],
        ["4", "第一行是示例数据，请参考填写"],
        ["5", "填写完成后请使用系统进行验证"],
    ]
    
    for row_idx, row_data in enumerate(info_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
            
            if row_idx == 1:
                cell.font = Font(name='微软雅黑', size=14, bold=True, color='4472C4')
            elif row_idx == 3:
                cell.font = Font(name='微软雅黑', size=11, bold=True)
                cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            elif row_idx == 14:
                cell.font = Font(name='微软雅黑', size=11, bold=True, color='FF0000')
            else:
                cell.font = Font(name='微软雅黑', size=10)
            
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 60
    
    # 保存文件
    wb.save(output_path)
    
    print(f"✓ 模板已生成: {output_path}")
    print(f"  - 数据行数: {num_rows}")
    print(f"  - 列数: {len(columns)}")
    print(f"  - 包含工作表: BOM数据、填写说明")
    print()
    
    return output_path


def main():
    """主函数"""
    import argparse
    
    parser = argparse.ArgumentParser(description='生成BOM申请模板')
    parser.add_argument('-o', '--output', help='输出文件路径')
    parser.add_argument('-n', '--num-rows', type=int, default=20, help='模板行数（默认20）')
    
    args = parser.parse_args()
    
    try:
        output_path = create_bom_template(args.output, args.num_rows)
        print("=" * 80)
        print("模板生成成功！")
        print("=" * 80)
        return 0
    except Exception as e:
        print(f"❌ 生成模板失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    sys.exit(main())
